from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, String, case, distinct
from typing import Optional, List
from datetime import date, datetime, timedelta
import pandas as pd
import io
import os
import tempfile
import shutil
import uuid
import openpyxl

from . import models, schemas, database, auth
from .database import SessionLocal, engine

models.Base.metadata.create_all(bind=engine)

# Автомиграции
with engine.connect() as _conn:
    for _sql in [
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS region VARCHAR(200)",
        """CREATE TABLE IF NOT EXISTS app_settings (
               key   VARCHAR(100) PRIMARY KEY,
               value VARCHAR(500)
           )""",
        "INSERT INTO app_settings (key, value) VALUES ('last_update', NULL) ON CONFLICT DO NOTHING",
        """CREATE TABLE IF NOT EXISTS login_logs (
               id         SERIAL PRIMARY KEY,
               username   VARCHAR(50),
               role       VARCHAR(20),
               region     VARCHAR(200),
               ip_address VARCHAR(64),
               user_agent VARCHAR(500),
               logged_at  TIMESTAMP DEFAULT NOW()
           )""",
        "CREATE INDEX IF NOT EXISTS idx_login_logs_logged_at ON login_logs (logged_at)",
        """CREATE TABLE IF NOT EXISTS stored_files (
               id            SERIAL PRIMARY KEY,
               original_name VARCHAR(300),
               stored_name   VARCHAR(300),
               size_bytes    BIGINT,
               content_type  VARCHAR(200),
               uploaded_by   VARCHAR(50),
               uploaded_at   TIMESTAMP DEFAULT NOW()
           )""",
    ]:
        try:
            _conn.execute(__import__('sqlalchemy').text(_sql))
            _conn.commit()
        except Exception:
            _conn.rollback()

app = FastAPI(title="Strahovka Insurance API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup_event():
    db = SessionLocal()
    try:
        auth.init_default_users(db)
    finally:
        db.close()


def insured_expr(M):
    """Застрахован: договор ещё не истёк и не расторгнут.

    Совпадает с эталонным запросом заказчика:
        date_end > <дата запроса> AND rescinding_date IS NULL
    Считается на дату запроса, а не на дату загрузки файла.
    """
    return and_(M.date_end != None, M.date_end > date.today(), M.rescinding_date == None)


def not_insured_expr(M):
    """Обратное к insured_expr, с явной обработкой пустой date_end."""
    return or_(M.date_end == None, M.date_end <= date.today(), M.rescinding_date != None)


# Госучреждения не обязаны страховать работников от несчастных случаев
STATE_INSTITUTION = "Государственное учреждение"


def active_insurance_expr(M):
    """Договор действует по новому скрипту заказчика (для метрики нарушителей):
        date_end > сегодня AND (rescinding_date IS NULL OR rescinding_date > сегодня)
    Отличие от insured_expr: расторжение будущей датой ещё считается активным.
    """
    return and_(
        M.date_end != None,
        M.date_end > date.today(),
        or_(M.rescinding_date == None, M.rescinding_date > date.today()),
    )


def is_insured_now(record) -> int:
    """То же правило, но для уже загруженной строки — колонка is_insured в БД
    хранит снимок на момент загрузки файла и для показа не годится."""
    return 1 if (
        record.date_end
        and record.date_end > date.today()
        and record.rescinding_date is None
    ) else 0


def apply_filters(query, model, params: dict, force_region: str = None):
    """Применяет все фильтры к запросу. force_region — обязательный регион для региональных пользователей."""
    M = model

    # Региональный пользователь — жёстко фиксируем регион, obl_name из params игнорируем
    if force_region:
        query = query.filter(M.obl_name == force_region)
        params = {k: v for k, v in params.items() if k != 'obl_name'}

    bin_ = params.get("bin")
    if bin_:
        # Убираем ведущие нули чтобы "000101651508" находил число 101651508
        bin_stripped = bin_.lstrip('0') or bin_
        query = query.filter(M.bin.cast(String).like(f"%{bin_stripped}%"))

    if params.get("bin_name"):
        query = query.filter(M.bin_name.ilike(f"%{params['bin_name']}%"))

    if params.get("system_delimiter_bin"):
        sdb = params['system_delimiter_bin'].lstrip('0') or params['system_delimiter_bin']
        query = query.filter(M.system_delimiter_bin.cast(String).like(f"%{sdb}%"))

    if params.get("system_delimiter_bin_name"):
        query = query.filter(M.system_delimiter_bin_name.ilike(f"%{params['system_delimiter_bin_name']}%"))

    if params.get("contract_number"):
        query = query.filter(M.contract_number.ilike(f"%{params['contract_number']}%"))

    if params.get("contract_date_from"):
        query = query.filter(M.contract_date >= params["contract_date_from"])
    if params.get("contract_date_to"):
        query = query.filter(M.contract_date <= params["contract_date_to"])

    if params.get("date_beg_from"):
        query = query.filter(M.date_beg >= params["date_beg_from"])
    if params.get("date_beg_to"):
        query = query.filter(M.date_beg <= params["date_beg_to"])

    if params.get("date_end_from") or params.get("date_end_to"):
        effective_end = case(
            (and_(M.rescinding_date != None, M.rescinding_date < M.date_end), M.rescinding_date),
            else_=M.date_end
        )
        if params.get("date_end_from"):
            query = query.filter(effective_end >= params["date_end_from"])
        if params.get("date_end_to"):
            query = query.filter(effective_end <= params["date_end_to"])

    if params.get("obl_name"):
        query = query.filter(M.obl_name.ilike(f"%{params['obl_name']}%"))
    if params.get("rai_name"):
        query = query.filter(M.rai_name.ilike(f"%{params['rai_name']}%"))
    if params.get("address"):
        query = query.filter(M.address.ilike(f"%{params['address']}%"))
    if params.get("phone"):
        query = query.filter(M.phone.ilike(f"%{params['phone']}%"))
    if params.get("leader_surname"):
        query = query.filter(M.leader_surname.ilike(f"%{params['leader_surname']}%"))
    if params.get("opf_name"):
        query = query.filter(M.opf_name.ilike(f"%{params['opf_name']}%"))
    if params.get("id_oked"):
        query = query.filter(M.id_oked.ilike(f"%{params['id_oked']}%"))
    if params.get("name_oked"):
        query = query.filter(M.name_oked.ilike(f"%{params['name_oked']}%"))

    if params.get("is_insured") is not None:
        query = query.filter(
            insured_expr(M) if params["is_insured"] == 1 else not_insured_expr(M)
        )

    if params.get("expires_in_months"):
        target_date = date.today() + timedelta(days=30 * params["expires_in_months"])
        query = query.filter(
            and_(
                M.date_end >= date.today(),
                M.date_end <= target_date
            )
        )

    return query


# ============ AUTH ============

# Системные учётки — не журналируем и не показываем в статистике входов
SYSTEM_ACCOUNTS = ("admin", "user")


@app.post("/api/auth/login", response_model=schemas.Token)
def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(database.get_db)):
    user = auth.authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Журналируем вход. Ошибка логирования не должна блокировать вход.
    if user.username not in SYSTEM_ACCOUNTS:
        try:
            fwd = request.headers.get("x-forwarded-for")
            ip = fwd.split(",")[0].strip() if fwd else (request.client.host if request.client else None)
            db.add(models.LoginLog(
                username=user.username,
                role=user.role,
                region=user.region,
                ip_address=ip,
                user_agent=(request.headers.get("user-agent") or "")[:500] or None,
                logged_at=datetime.now(),
            ))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"Failed to write login log: {e}")

    access_token = auth.create_access_token(
        data={"sub": user.username, "role": user.role}
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/api/auth/me", response_model=schemas.UserResponse)
def get_me(current_user: models.User = Depends(auth.get_current_active_user)):
    return current_user


# ============ METRICS ============

@app.get("/api/metrics", response_model=schemas.MetricsResponse)
def get_metrics(
    bin: Optional[str] = None,
    bin_name: Optional[str] = None,
    system_delimiter_bin: Optional[str] = None,
    system_delimiter_bin_name: Optional[str] = None,
    contract_number: Optional[str] = None,
    contract_date_from: Optional[date] = None,
    contract_date_to: Optional[date] = None,
    date_beg_from: Optional[date] = None,
    date_beg_to: Optional[date] = None,
    date_end_from: Optional[date] = None,
    date_end_to: Optional[date] = None,
    obl_name: Optional[str] = None,
    rai_name: Optional[str] = None,
    address: Optional[str] = None,
    phone: Optional[str] = None,
    leader_surname: Optional[str] = None,
    opf_name: Optional[str] = None,
    id_oked: Optional[str] = None,
    name_oked: Optional[str] = None,
    is_insured: Optional[int] = None,
    expires_in_months: Optional[int] = None,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    params = locals()
    params.pop("current_user"); params.pop("db")
    M = models.InsuranceRecord

    # Карточки «Всего» и «Застрахованы»: строки договоров + уникальные БИН.
    # case(...) без else_ даёт NULL, а count(distinct) его пропускает.
    query = apply_filters(
        db.query(
            func.count().label("total"),
            func.sum(case((insured_expr(M), 1), else_=0)).label("insured"),
            func.count(distinct(M.bin)).label("total_bins"),
            func.count(distinct(case((insured_expr(M), M.bin)))).label("insured_bins"),
        ),
        M,
        params,
        force_region=current_user.region,
    )
    row = query.one()
    total = row.total or 0
    insured = row.insured or 0

    # Карточка «Не застрахованы»: нарушители — компании (БИН), обязанные
    # страховать работников, но без действующего договора.
    #   обязан = не госучреждение И esutd_akt_td >= 2
    #   не застрахован = ни одной строки с активной страховкой
    # Поля паспорта (opf_name, esutd) в выгрузке одинаковы для всех строк БИН,
    # поэтому берём их через max().
    per_bin = apply_filters(
        db.query(
            M.bin.label("bin"),
            func.max(M.esutd_akt_td).label("esutd"),
            func.max(M.opf_name).label("opf"),
            func.sum(case((active_insurance_expr(M), 1), else_=0)).label("active_cnt"),
        ),
        M,
        params,
        force_region=current_user.region,
    ).group_by(M.bin).subquery()

    eligible_filter = and_(
        per_bin.c.esutd >= 2,
        per_bin.c.opf != STATE_INSTITUTION,  # != исключает и NULL
    )
    row2 = db.query(
        func.count().label("eligible"),
        func.sum(case((per_bin.c.active_cnt == 0, 1), else_=0)).label("violators"),
    ).select_from(per_bin).filter(eligible_filter).one()

    return schemas.MetricsResponse(
        total=total,
        insured=insured,
        not_insured=total - insured,
        total_bins=row.total_bins or 0,
        insured_bins=row.insured_bins or 0,
        violators=row2.violators or 0,
        eligible_total=row2.eligible or 0,
    )


# ============ RECORDS ============

@app.get("/api/records", response_model=schemas.InsuranceRecordList)
def get_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
    sort_by: Optional[str] = "id",
    sort_order: Optional[str] = "asc",
    bin: Optional[str] = None,
    bin_name: Optional[str] = None,
    system_delimiter_bin: Optional[str] = None,
    system_delimiter_bin_name: Optional[str] = None,
    contract_number: Optional[str] = None,
    contract_date_from: Optional[date] = None,
    contract_date_to: Optional[date] = None,
    date_beg_from: Optional[date] = None,
    date_beg_to: Optional[date] = None,
    date_end_from: Optional[date] = None,
    date_end_to: Optional[date] = None,
    obl_name: Optional[str] = None,
    rai_name: Optional[str] = None,
    address: Optional[str] = None,
    phone: Optional[str] = None,
    leader_surname: Optional[str] = None,
    opf_name: Optional[str] = None,
    id_oked: Optional[str] = None,
    name_oked: Optional[str] = None,
    is_insured: Optional[int] = None,
    expires_in_months: Optional[int] = None,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    params = {k: v for k, v in locals().items() if k not in ("page", "page_size", "sort_by", "sort_order", "current_user", "db")}
    query = apply_filters(db.query(models.InsuranceRecord), models.InsuranceRecord, params, force_region=current_user.region)

    total = query.count()

    if sort_by and hasattr(models.InsuranceRecord, sort_by):
        order_col = getattr(models.InsuranceRecord, sort_by)
        query = query.order_by(order_col.desc() if sort_order == "desc" else order_col.asc())

    records = query.offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for r in records:
        item = schemas.InsuranceRecordResponse.model_validate(r)
        item.is_insured = is_insured_now(r)
        items.append(item)

    return schemas.InsuranceRecordList(items=items, total=total, page=page, page_size=page_size)


@app.get("/api/records/download")
def download_records(
    bin: Optional[str] = None,
    bin_name: Optional[str] = None,
    system_delimiter_bin: Optional[str] = None,
    system_delimiter_bin_name: Optional[str] = None,
    contract_number: Optional[str] = None,
    contract_date_from: Optional[date] = None,
    contract_date_to: Optional[date] = None,
    date_beg_from: Optional[date] = None,
    date_beg_to: Optional[date] = None,
    date_end_from: Optional[date] = None,
    date_end_to: Optional[date] = None,
    obl_name: Optional[str] = None,
    rai_name: Optional[str] = None,
    address: Optional[str] = None,
    phone: Optional[str] = None,
    leader_surname: Optional[str] = None,
    opf_name: Optional[str] = None,
    id_oked: Optional[str] = None,
    name_oked: Optional[str] = None,
    is_insured: Optional[int] = None,
    expires_in_months: Optional[int] = None,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    params = {k: v for k, v in locals().items() if k not in ("current_user", "db")}
    query = apply_filters(db.query(models.InsuranceRecord), models.InsuranceRecord, params, force_region=current_user.region)
    records = query.all()

    data = []
    for r in records:
        data.append({
            'БИН': str(int(r.bin)).zfill(12) if r.bin is not None else '',
            'Название компании': r.bin_name,
            'БИН страховой компании': str(int(r.system_delimiter_bin)).zfill(12) if r.system_delimiter_bin is not None else '',
            'Страховая компания': r.system_delimiter_bin_name,
            'Номер договора': r.contract_number,
            'Дата договора': r.contract_date,
            'Дата начала': r.date_beg,
            'Дата окончания': r.date_end,
            'Дата расторжения': r.rescinding_date,
            'Сумма': r.calculated_amount,
            'Застрахованных сотр.': r.count_employees,
            'Всего сотрудников': r.total_employees_count,
            'Кол-во 12 мес.': r.kol_12mes,
            'ФОТ 12 мес.': r.fot_12mes,
            'ESUTD акт. ТД': r.esutd_akt_td,
            'Область': r.obl_name,
            'Район': r.rai_name,
            'Адрес': r.address,
            'Телефон': r.phone,
            'Руководитель': f"{r.leader_surname or ''} {r.leader_name or ''} {r.leader_middlename or ''}".strip(),
            'ОПФ': r.opf_name,
            'Код ОКЭД': r.id_oked,
            'Вид деятельности (ОКЭД)': r.name_oked,
            'ИП': r.ip,
            'Флаг': r.flag_head,
            'Застрахован': 'Да' if is_insured_now(r) else 'Нет',
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl', datetime_format='DD.MM.YYYY', date_format='DD.MM.YYYY') as writer:
        df.to_excel(writer, index=False, sheet_name='Insurance Records')
        ws = writer.sheets['Insurance Records']
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 50)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=insurance_records.xlsx"}
    )


# ============ UPLOAD (ADMIN ONLY) ============

@app.post("/api/upload")
def upload_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only Excel files are allowed")

    tmp_path = None
    try:
        # Stream to disk — не грузим 145MB целиком в RAM
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name

        # read_only=True — потоковое чтение, память ~200MB вместо 12GB
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter))

        def safe_int(val):
            if val is None:
                return None
            try:
                return int(val)
            except Exception:
                return None

        def safe_float(val):
            if val is None:
                return None
            try:
                f = float(val)
                return None if f != f else f
            except Exception:
                return None

        def safe_str(val):
            return str(val) if val is not None else None

        def safe_date(val):
            if val is None:
                return None
            if isinstance(val, str):
                return None  # malformed string dates (e.g. '16.04.0212') → NULL
            return val

        db.query(models.InsuranceRecord).delete()
        db.commit()

        now = datetime.now()
        records = []
        BATCH = 5000

        for idx, row_vals in enumerate(rows_iter):
            try:
                row = dict(zip(headers, row_vals))
                date_end_val = safe_date(row.get('DATE_END'))
                rescinding_val = safe_date(row.get('RESCINDING_DATE'))

                # Договор действует до даты окончания, но расторжение обрывает его раньше
                effective_end = date_end_val
                if rescinding_val and (not date_end_val or rescinding_val < date_end_val):
                    effective_end = rescinding_val

                records.append({
                    'row_num': safe_int(row.get('   ', idx + 1)),
                    'bin': safe_int(row.get('BIN')),
                    'system_delimiter_bin': safe_float(row.get('SYSTEM_DELIMITER_BIN')),
                    'system_delimiter_bin_name': safe_str(row.get('SYSTEM_DELIMITER_BIN_NAME')),
                    'contract_number': safe_str(row.get('CONTRACT_NUMBER')),
                    'contract_date': safe_date(row.get('CONTRACT_DATE')),
                    'date_beg': safe_date(row.get('DATE_BEG')),
                    'date_end': date_end_val,
                    'rescinding_date': rescinding_val,
                    'calculated_amount': safe_float(row.get('CALCULATED_AMOUNT')),
                    'count_employees': safe_float(row.get('COUNT_EMPLOYEES')),
                    'total_employees_count': safe_float(row.get('TOTAL_EMPLOYEES_COUNT')),
                    'id_system': safe_float(row.get('ID')),
                    'flag_head': safe_float(row.get('FLAG_HEAD')),
                    'sys_date': safe_date(row.get('SYS_DATE')),
                    'bin_name': safe_str(row.get('BIN_NAME')),
                    'id_reg': safe_int(row.get('ID_REG')),
                    'obl_name': safe_str(row.get('OBL_NAME')),
                    'idrai': safe_int(row.get('IDRAI')),
                    'rai_name': safe_str(row.get('RAI_NAME')),
                    'address': safe_str(row.get('ADDRESS')),
                    'phone': safe_str(row.get('PHONE')),
                    'mail': safe_float(row.get('MAIL')),
                    'leader_surname': safe_str(row.get('LEADER_SURNAME')),
                    'leader_name': safe_str(row.get('LEADER_NAME')),
                    'leader_middlename': safe_str(row.get('LEADER_MIDDLENAME')),
                    'opf': safe_float(row.get('OPF')),
                    'opf_name': safe_str(row.get('OPF_NAME')),
                    'id_oked': safe_str(row.get('ID_OKED')),
                    'name_oked': safe_str(row.get('NAME_OKED')),
                    'kol_12mes': safe_int(row.get('KOL_12MES')),
                    'fot_12mes': safe_int(row.get('FOT_12MES')),
                    'esutd_akt_td': safe_int(row.get('ESUTD_AKT_TD')),
                    'ip': safe_int(row.get('IP')),
                    'tip': safe_int(row.get('TIP')),
                    'is_insured': 1 if (effective_end and effective_end > now) else 0,
                    'created_at': now,
                    'updated_at': now,
                })

                if len(records) >= BATCH:
                    try:
                        db.bulk_insert_mappings(models.InsuranceRecord, records)
                        db.commit()
                    except Exception as batch_err:
                        db.rollback()
                        print(f"Batch insert failed: {batch_err}")
                    records = []

            except Exception as e:
                print(f"Error processing row {idx}: {e}")
                continue

        if records:
            try:
                db.bulk_insert_mappings(models.InsuranceRecord, records)
                db.commit()
            except Exception as batch_err:
                db.rollback()
                print(f"Final batch insert failed: {batch_err}")

        wb.close()

        count = db.query(models.InsuranceRecord).count()
        today = date.today().strftime("%d.%m.%Y")
        setting = db.query(models.AppSetting).filter(models.AppSetting.key == "last_update").first()
        if setting:
            setting.value = today
        else:
            db.add(models.AppSetting(key="last_update", value=today))
        db.commit()
        return {"message": f"Successfully uploaded {count} records"}

    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error processing file: {str(e)}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ============ REGIONS & DISTRICTS ============

@app.get("/api/regions")
def get_regions(
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """Список уникальных областей из таблицы (для фильтра)"""
    rows = (
        db.query(models.InsuranceRecord.obl_name)
        .filter(models.InsuranceRecord.obl_name.isnot(None))
        .distinct()
        .order_by(models.InsuranceRecord.obl_name)
        .all()
    )
    return [r[0] for r in rows if r[0]]


@app.get("/api/districts")
def get_districts(
    region: str,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """Список уникальных районов для заданной области"""
    rows = (
        db.query(models.InsuranceRecord.rai_name)
        .filter(
            models.InsuranceRecord.obl_name == region,
            models.InsuranceRecord.rai_name.isnot(None),
        )
        .distinct()
        .order_by(models.InsuranceRecord.rai_name)
        .all()
    )
    return [r[0] for r in rows if r[0]]


# ============ USER MANAGEMENT (ADMIN) ============

@app.get("/api/users", response_model=List[schemas.UserResponse])
def list_users(
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    return db.query(models.User).order_by(models.User.id).all()


@app.post("/api/users", response_model=schemas.UserResponse)
def create_user(
    data: schemas.UserCreate,
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    if db.query(models.User).filter(models.User.username == data.username).first():
        raise HTTPException(400, "Пользователь с таким логином уже существует")
    user = models.User(
        username=data.username,
        hashed_password=auth.get_password_hash(data.password),
        role=data.role,
        region=data.region,
        is_active=1,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@app.delete("/api/users/{user_id}")
def delete_user(
    user_id: int,
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    if user.username in ("admin", "user"):
        raise HTTPException(400, "Нельзя удалить системных пользователей")
    db.delete(user)
    db.commit()
    return {"message": "Удалён"}


# ============ LOGIN LOGS ============

@app.get("/api/logs", response_model=schemas.LoginLogList)
def get_login_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    username: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    query = db.query(models.LoginLog).filter(
        models.LoginLog.username.notin_(SYSTEM_ACCOUNTS)
    )

    if username:
        query = query.filter(models.LoginLog.username.ilike(f"%{username}%"))
    if date_from:
        query = query.filter(models.LoginLog.logged_at >= date_from)
    if date_to:
        # включаем весь день date_to
        query = query.filter(models.LoginLog.logged_at < date_to + timedelta(days=1))

    total = query.count()
    items = (
        query.order_by(models.LoginLog.logged_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return schemas.LoginLogList(items=items, total=total, page=page, page_size=page_size)


NO_REGION_LABEL = "АО КСЖ ГАК"

_MONTHS_GEN = (
    "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
)


def _period_caption(date_from: Optional[date], date_to: Optional[date]) -> str:
    """Подпись колонки: «за период с 8 – 9 июля 2026 года» и т.п."""
    if not date_from and not date_to:
        return "Кол-во авторизаций за весь период"

    def full(d):
        return f"{d.day} {_MONTHS_GEN[d.month - 1]} {d.year} года"

    if date_from and date_to:
        if (date_from.month, date_from.year) == (date_to.month, date_to.year):
            return (f"Кол-во авторизаций за период с {date_from.day} – "
                    f"{date_to.day} {_MONTHS_GEN[date_to.month - 1]} {date_to.year} года")
        return f"Кол-во авторизаций за период с {full(date_from)} по {full(date_to)}"
    if date_from:
        return f"Кол-во авторизаций с {full(date_from)}"
    return f"Кол-во авторизаций по {full(date_to)}"


@app.get("/api/logs/export")
def export_login_stats(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    """Excel: статистика авторизаций по регионам. Без дат — за всё время."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    L = models.LoginLog

    # Все учётные записи, включая тех, кто ни разу не заходил
    users = (
        db.query(models.User)
        .filter(models.User.username.notin_(SYSTEM_ACCOUNTS))
        .all()
    )

    stats_q = db.query(
        L.username,
        func.count().label("logins"),
        func.max(L.logged_at).label("last_login"),
    ).filter(L.username.notin_(SYSTEM_ACCOUNTS))

    if date_from:
        stats_q = stats_q.filter(L.logged_at >= date_from)
    if date_to:
        stats_q = stats_q.filter(L.logged_at < date_to + timedelta(days=1))

    stats = {r.username: r for r in stats_q.group_by(L.username).all()}

    # Группируем учётки по регионам
    regions = {}
    for u in users:
        regions.setdefault(u.region or NO_REGION_LABEL, []).append(u)

    # Регионы по алфавиту, «АО КСЖ ГАК» — в конец
    ordered = sorted(regions, key=lambda r: (r == NO_REGION_LABEL, r.lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "Статистика входа пользователей АО «КСЖ «ГАК» в портал"
    title.font = Font(bold=True, size=12)
    title.alignment = center

    headers = ["№", "Регион", "Кол-во учетных записей",
               _period_caption(date_from, date_to),
               "Дата и время авторизации последней авторизации"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=name)
        c.font = Font(bold=True)
        c.alignment = left if col == 2 else center
        c.border = border

    row = 3
    for num, region in enumerate(ordered, start=1):
        accounts = regions[region]
        # Активные сверху, самый свежий вход первым
        accounts.sort(
            key=lambda u: (stats[u.username].last_login if u.username in stats else datetime.min),
            reverse=True,
        )
        total_logins = sum(stats[u.username].logins for u in accounts if u.username in stats)
        first_row, last_row = row, row + len(accounts) - 1

        for u in accounts:
            last_login = stats[u.username].last_login if u.username in stats else None
            c = ws.cell(row=row, column=5)
            if last_login:
                c.value = last_login
                c.number_format = "DD.MM.YYYY, HH:MM:SS"
            c.alignment = left
            c.border = border
            for col in (1, 2, 3, 4):
                ws.cell(row=row, column=col).border = border
            row += 1

        for col, value, align in (
            (1, num, center),
            (2, region, left),
            (3, len(accounts), center),
            (4, total_logins, center),
        ):
            if last_row > first_row:
                ws.merge_cells(start_row=first_row, start_column=col,
                               end_row=last_row, end_column=col)
            cell = ws.cell(row=first_row, column=col, value=value)
            cell.alignment = align
            cell.border = border

    for col, width in enumerate((6, 34, 20, 24, 30), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 32

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"logins_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ============ SHARED FILES ============

FILES_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads", "files")
os.makedirs(FILES_DIR, exist_ok=True)

MAX_FILE_SIZE = 200 * 1024 * 1024  # 200 МБ — столько же пропускает nginx


@app.get("/api/files", response_model=List[schemas.StoredFileResponse])
def list_files(
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    return db.query(models.StoredFile).order_by(models.StoredFile.uploaded_at.desc()).all()


@app.post("/api/files", response_model=schemas.StoredFileResponse)
def upload_shared_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    if not file.filename:
        raise HTTPException(400, "Файл без имени")

    ext = os.path.splitext(file.filename)[1][:20]
    stored_name = f"{uuid.uuid4().hex}{ext}"
    dest = os.path.join(FILES_DIR, stored_name)

    size = 0
    try:
        with open(dest, "wb") as out:
            while chunk := file.file.read(1024 * 1024):
                size += len(chunk)
                if size > MAX_FILE_SIZE:
                    raise HTTPException(413, "Файл больше 200 МБ")
                out.write(chunk)
    except Exception:
        if os.path.exists(dest):
            os.unlink(dest)
        raise

    record = models.StoredFile(
        original_name=os.path.basename(file.filename)[:300],
        stored_name=stored_name,
        size_bytes=size,
        content_type=file.content_type,
        uploaded_by=current_user.username,
        uploaded_at=datetime.now(),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@app.get("/api/files/{file_id}/download")
def download_shared_file(
    file_id: int,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    record = db.query(models.StoredFile).filter(models.StoredFile.id == file_id).first()
    if not record:
        raise HTTPException(404, "Файл не найден")

    path = os.path.join(FILES_DIR, record.stored_name)
    if not os.path.exists(path):
        raise HTTPException(404, "Файл отсутствует на диске")

    return FileResponse(
        path,
        filename=record.original_name,
        media_type=record.content_type or "application/octet-stream",
    )


@app.delete("/api/files/{file_id}")
def delete_shared_file(
    file_id: int,
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    record = db.query(models.StoredFile).filter(models.StoredFile.id == file_id).first()
    if not record:
        raise HTTPException(404, "Файл не найден")

    path = os.path.join(FILES_DIR, record.stored_name)
    if os.path.exists(path):
        os.unlink(path)
    db.delete(record)
    db.commit()
    return {"message": "Удалён"}


# ============ LAST UPDATE DATE ============

@app.get("/api/settings/last_update")
def get_last_update(
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    row = db.query(models.AppSetting).filter(models.AppSetting.key == "last_update").first()
    return {"last_update": row.value if row else None}


@app.put("/api/settings/last_update")
def set_last_update(
    body: dict,
    current_user: models.User = Depends(auth.require_admin),
    db: Session = Depends(database.get_db)
):
    value = body.get("last_update", "")
    row = db.query(models.AppSetting).filter(models.AppSetting.key == "last_update").first()
    if row:
        row.value = value
    else:
        db.add(models.AppSetting(key="last_update", value=value))
    db.commit()
    return {"last_update": value}


@app.get("/api/suggestions")
def get_suggestions(
    field: str,
    query: str = "",
    limit: int = Query(10, le=20),
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    allowed = {
        'bin': models.InsuranceRecord.bin,
        'bin_name': models.InsuranceRecord.bin_name,
        'system_delimiter_bin': models.InsuranceRecord.system_delimiter_bin,
        'system_delimiter_bin_name': models.InsuranceRecord.system_delimiter_bin_name,
    }
    if field not in allowed:
        raise HTTPException(400, "Invalid field")

    bin_fields = {'bin', 'system_delimiter_bin'}
    col = allowed[field]
    q = db.query(col).filter(col.isnot(None))
    if query:
        # Для БИН-полей убираем ведущие нули перед сравнением
        search = (query.lstrip('0') or query) if field in bin_fields else query
        q = q.filter(col.cast(String).ilike(f"%{search}%"))
    rows = q.distinct().limit(limit).all()
    # Для БИН-полей форматируем с нулями в ответе
    if field in bin_fields:
        return [str(int(float(r[0]))).zfill(12) for r in rows if r[0] is not None]
    return [str(r[0]) for r in rows if r[0] is not None]


@app.get("/api/health")
def health_check():
    return {"status": "ok"}
